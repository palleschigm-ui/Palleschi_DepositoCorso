# ESEMPI DI INPUT/OUTPUT (I/O) SU FILE IN PYTHON
# Formati: TXT - CSV - EXCEL (.xlsx)
import os
print("Sto leggendo da:", os.getcwd())
os.chdir("C:/Users/Gianmario/Desktop/Python/Esercitazioni/Corso_GiGroup/Palleschi_DepositoCorso/Lezione_08/lettura_file")

# ============================================================
# 1) I/O SU FILE DI TESTO (.txt)
# ============================================================

# --------------------------------------------
# SCRITTURA SU FILE TXT
# --------------------------------------------
def scrivi_txt():
    # 'with' apre il file e garantisce la chiusura automatica
    with open("dati.txt", "w", encoding="utf-8") as file:
        # .write() scrive una stringa sul file
        file.write("Ciao!\n")
        file.write("Questa è una riga salvata su un file di testo.\n")
        file.write("Python rende facile leggere e scrivere file.")

    print("File TXT scritto correttamente.")


# --------------------------------------------
# LETTURA SU FILE TXT (read completo)
# --------------------------------------------
def leggi_txt():
    # modalità "r" = read
    with open("dati.txt", "r", encoding="utf-8") as file:
        # .read() legge tutto il contenuto del file come un'unica stringa
        contenuto = file.read()
    print("CONTENUTO DEL FILE TXT:")
    print(contenuto)


# --------------------------------------------
# LETTURA RIGA PER RIGA
# --------------------------------------------
def leggi_righe_txt():
    # Il file è un iteratore: si può ciclare direttamente sulle righe
    with open("dati.txt", "r", encoding="utf-8") as file:
        for riga in file:
            # strip() rimuove \n e spazi superflui
            print("Riga:", riga.strip())


# ============================================================
# 2) I/O SU FILE CSV (.csv)
# ============================================================

# Il modulo csv è nella libreria standard di Python
import csv

# --------------------------------------------
# SCRITTURA CSV
# --------------------------------------------
def scrivi_csv():
    # newline="" evita righe vuote inutili su Windows
    with open("dati.csv", "w", newline="", encoding="utf-8") as file:
        # writer gestisce la scrittura riga per riga
        writer = csv.writer(file) #prepara la formattazione delle righe CSV

        # intestazione (prima riga del CSV)
        writer.writerow(["Nome", "Età", "Città"]) #aggiunge una riga nel file

        # righe di dati
        writer.writerow(["Mirko", 30, "Torino"])
        writer.writerow(["Anna", 25, "Roma"])
        writer.writerow(["Giuseppe", 28, "Napoli"])

    print("File CSV scritto correttamente.")


# --------------------------------------------
# LETTURA CSV
# --------------------------------------------
def leggi_csv():
    print("CONTENUTO DEL FILE CSV:")
    with open("dati.csv", "r", encoding="utf-8") as file:
        # reader converte ogni riga in una lista di valori
        reader = csv.reader(file)

        for riga in reader:
            print(riga)  # ogni riga è una lista di stringhe


# ============================================================
# 3) I/O SU FILE EXCEL (.xlsx)
# ============================================================

# Per Excel serve installare openpyxl:
# pip install openpyxl
from openpyxl import Workbook, load_workbook

# --------------------------------------------
# SCRITTURA SU EXCEL
# --------------------------------------------
def scrivi_excel():
    wb = Workbook()     # crea un nuovo file Excel
    ws = wb.active      # ottiene il foglio attivo
    ws.title = "Dati"   # rinomina il foglio

    # append() inserisce una riga alla volta
    ws.append(["Nome", "Età", "Città"])
    ws.append(["Mirko", 30, "Torino"])
    ws.append(["Anna", 25, "Roma"])
    ws.append(["Giuseppe", 28, "Napoli"])

    wb.save("dati.xlsx")   # salva il file Excel
    print("File Excel scritto correttamente.")


# --------------------------------------------
# LETTURA SU EXCEL
# --------------------------------------------
def leggi_excel():
    wb = load_workbook("dati.xlsx")   # apre il file Excel
    ws = wb["Dati"]                   # seleziona il foglio per nome

    print("CONTENUTO FILE EXCEL:")
    # iter_rows restituisce una riga per volta
    # values_only=True → ottieni solo i valori, non gli oggetti delle celle
    for row in ws.iter_rows(values_only=True):
        print(row)


# ============================================================
# MAIN DI TEST
# ============================================================

if __name__ == "__main__":
    print("\n--- SCRITTURA TXT ---")
    scrivi_txt()

    print("\n--- LETTURA TXT ---")
    leggi_txt()

    print("\n--- LETTURA RIGA PER RIGA (TXT) ---")
    leggi_righe_txt()

    print("\n--- SCRITTURA CSV ---")
    scrivi_csv()

    print("\n--- LETTURA CSV ---")
    leggi_csv()

    print("\n--- SCRITTURA EXCEL ---")
    scrivi_excel()

    print("\n--- LETTURA EXCEL ---")
    leggi_excel()
